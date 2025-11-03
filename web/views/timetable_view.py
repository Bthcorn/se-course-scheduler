"""
Timetable View page for the Streamlit application
"""

from datetime import datetime
from typing import Any
import io
import re
import streamlit as st
import streamlit_shadcn_ui as ui
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from web.utils import extract_unique_values
from web.components import display_schedule_table
from web.views.base_page import BasePage


class TimetableViewPage(BasePage):
    """Timetable View page class"""

    def format_cell_content(
        self, course_name: str, professor: str, room: str, is_reserved: bool = False
    ) -> str:
        """Format cell content with HTML styling"""
        if is_reserved:
            return (
                f"<div style='text-align: center;'>"
                f"<strong style='color: #1976d2; font-size: 1.1em;'>[RESERVED]</strong><br/>"
                f"<span style='color: #666;'>{course_name}</span><br/>"
                f"<span style='color: #999; font-size: 0.9em;'>Room: {room}</span>"
                f"</div>"
            )
        else:
            return (
                f"<div style='text-align: center;'>"
                f"<strong style='color: #1f77b4; font-size: 1.1em; font-weight: 600;'>"
                f"{course_name}</strong><br/>"
                f"<span style='color: #ff6b35; font-style: italic; font-weight: 500;'>"
                f"{professor}</span><br/>"
                f"<span style='color: #666; font-size: 0.85em;'>Room: {room}</span>"
                f"</div>"
            )

    @staticmethod
    def parse_html_cell(html_content: str) -> dict[str, Any]:
        """Parse HTML cell content to extract course, professor, room"""
        # Extract course name (bold text in blue)
        course_match = re.search(
            r"<strong[^>]*>(.*?)</strong>", html_content, re.DOTALL
        )
        course = course_match.group(1).strip() if course_match else ""

        # Extract professor (italic/orange text)
        prof_match = re.search(
            r"<span[^>]*color: #ff6b35[^>]*>(.*?)</span>", html_content, re.DOTALL
        )
        professor = prof_match.group(1).strip() if prof_match else ""

        # Extract room (small gray text)
        room_match = re.search(r"Room: (\w+)", html_content)
        room = room_match.group(1) if room_match else ""

        # Check if reserved
        is_reserved = "[RESERVED]" in html_content or "RESERVED" in html_content.upper()

        return {
            "course": course,
            "professor": professor,
            "room": room,
            "is_reserved": is_reserved,
            "raw_text": re.sub("<[^<]+?>", "", html_content).replace("\n", " ").strip(),
        }

    def create_timetable_data(
        self,
        schedule: list[dict[str, str]],
        reserved: list[dict[str, str]],
        room_filter: str,
        professor_filter: str,
        year_filter: str = "All",
    ) -> dict[str, dict[str, dict[str, list[str]]]]:
        """Create timetable data structure: {room: {day: {period: [items]}}}"""
        days_order = [
            "monday",
            "tuesday",
            "wednesday",
            "thursday",
            "friday",
            "saturday",
            "sunday",
        ]
        periods_order = ["morning", "afternoon", "evening"]

        # Group by room
        timetables_by_room: dict[str, dict[str, dict[str, list[str]]]] = {}

        # Helper function to initialize timetable
        def init_timetable():
            timetable: dict[str, dict[str, list[str]]] = {}
            for day in days_order:
                timetable[day] = {}
                for period in periods_order:
                    timetable[day][period] = []
            return timetable

        # Add scheduled courses
        for item in schedule:
            room = item.get("room", "")
            professor = item.get("professor", "")
            year = item.get("year", "")

            # Apply filters
            if room_filter != "All" and room != room_filter:
                continue
            if professor_filter != "All" and professor != professor_filter:
                continue
            if year_filter != "All" and year != year_filter:
                continue

            if room not in timetables_by_room:
                timetables_by_room[room] = init_timetable()

            day = item.get("day", "").lower()
            period = item.get("period", "").lower()

            if (
                day in timetables_by_room[room]
                and period in timetables_by_room[room][day]
            ):
                course_name = item.get("course_name", "")
                cell_content = self.format_cell_content(
                    course_name, professor, room, False
                )
                timetables_by_room[room][day][period].append(cell_content)

        # Add reserved slots
        for item in reserved:
            room = item.get("room", "")
            reason = item.get("reason", "Reserved")

            # Apply filters
            if room_filter != "All" and room != room_filter:
                continue

            if room not in timetables_by_room:
                timetables_by_room[room] = init_timetable()

            day = item.get("day", "").lower()
            period = item.get("period", "").lower()

            if (
                day in timetables_by_room[room]
                and period in timetables_by_room[room][day]
            ):
                cell_content = self.format_cell_content(reason, "", room, True)
                timetables_by_room[room][day][period].append(cell_content)

        return timetables_by_room

    @staticmethod
    def timetable_to_dataframe(
        timetable: dict[str, dict[str, list[str]]],
    ) -> pd.DataFrame:
        """Convert timetable data to DataFrame"""
        days_order = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]
        periods_order = ["Morning", "Afternoon", "Evening"]

        table_data = []
        for day in days_order:
            day_lower = day.lower()
            row = {"Day": day}
            for period in periods_order:
                period_lower = period.lower()
                items = timetable.get(day_lower, {}).get(period_lower, [])
                if items:
                    # Extract text from HTML, removing HTML tags for Excel
                    text_items = []
                    for item in items:
                        # Remove HTML tags and get plain text
                        text = re.sub("<[^<]+?>", "", item)
                        text = text.replace("\n", " ").strip()
                        text_items.append(text)
                    row[period] = " / ".join(text_items)
                else:
                    row[period] = "Available"
            table_data.append(row)

        df = pd.DataFrame(table_data)
        df = df.set_index("Day")
        return df

    def export_timetable_to_excel(
        self, room: str, timetable: dict[str, dict[str, list[str]]]
    ) -> bytes:
        """Export timetable to Excel with formatting matching the web view"""

        days_order = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]
        periods_order = ["Morning", "Afternoon", "Evening"]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Room_{room.upper()}"

        # Header row with dark blue background
        ws.merge_cells(f"A1:{get_column_letter(len(periods_order) + 1)}1")
        ws["A1"] = f"Room: {room.upper()}"
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
        ws["A1"].fill = PatternFill(
            start_color="1f77b4", end_color="1f77b4", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Column headers
        headers = ["Day"] + periods_order
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            cell_ref = f"{col_letter}2"
            ws[cell_ref] = header
            cell = ws[cell_ref]
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(
                start_color="f0f2f6", end_color="f0f2f6", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center" if col_idx > 1 else "left", vertical="center"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Data rows
        for row_idx, day in enumerate(days_order, start=3):
            day_lower = day.lower()

            # Day column
            day_cell_ref = f"A{row_idx}"
            ws[day_cell_ref] = day
            day_cell = ws[day_cell_ref]
            day_cell.font = Font(bold=True, size=10)
            day_cell.fill = PatternFill(
                start_color="f9f9f9", end_color="f9f9f9", fill_type="solid"
            )
            day_cell.alignment = Alignment(horizontal="left", vertical="center")
            day_cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Period columns
            for col_idx, period in enumerate(periods_order, start=2):
                period_lower = period.lower()
                items = timetable.get(day_lower, {}).get(period_lower, [])

                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_idx}"
                cell = ws[cell_ref]
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                if items:
                    # Parse first item (usually only one per cell)
                    parsed = TimetableViewPage.parse_html_cell(items[0])

                    if parsed["is_reserved"]:
                        cell.fill = PatternFill(
                            start_color="e3f2fd", end_color="e3f2fd", fill_type="solid"
                        )
                        cell.value = parsed["raw_text"]
                        cell.font = Font(size=9, color="1976d2", bold=True)
                    else:
                        cell.fill = PatternFill(
                            start_color="f1f8e9", end_color="f1f8e9", fill_type="solid"
                        )
                        # Create formatted text (course name, professor, room)
                        lines = []
                        if parsed["course"]:
                            lines.append(parsed["course"])
                        if parsed["professor"]:
                            lines.append(parsed["professor"])
                        if parsed["room"]:
                            lines.append(f"Room: {parsed['room']}")
                        cell.value = "\n".join(lines)
                        cell.font = Font(size=9)
                else:
                    cell.value = "Available"
                    cell.fill = PatternFill(
                        start_color="fafafa", end_color="fafafa", fill_type="solid"
                    )
                    cell.font = Font(size=9, color="999999", italic=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 30
        ws.row_dimensions[1].height = 30
        for row in range(3, len(days_order) + 3):
            ws.row_dimensions[row].height = 80

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def convert_period_to_time(self, period: str) -> str:
        """Convert period to time"""
        if period == "Morning":
            return "Morning (9:00-12:00)"
        elif period == "Afternoon":
            return "Afternoon (13:00-16:00)"
        elif period == "Evening":
            return "Evening (17:00-20:00)"
        return ""

    def render_timetable_for_room(
        self, room: str, timetable: dict[str, dict[str, list[str]]]
    ):
        """Render a timetable table for a specific room"""
        days_order = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]
        periods_order = ["Morning", "Afternoon", "Evening"]

        # Create HTML table
        html_table = "<table style='width:100%; border-collapse: collapse; margin-bottom: 2rem;'>"
        html_table += "<thead><tr style='background-color: #1f77b4; color: white;'>"
        html_table += (
            f"<th colspan='4' style='padding: 12px; border: 1px solid #ddd; "
            f"text-align: center; font-size: 1.2em;'>"
            f"Room: {room.upper()}</th>"
        )
        html_table += "</tr><tr style='background-color: #f0f2f6;'>"
        html_table += (
            "<th style='padding: 10px; border: 1px solid #ddd; text-align: left; "
            "font-weight: 600;'>Day</th>"
        )
        for period in periods_order:
            html_table += (
                f"<th style='padding: 10px; border: 1px solid #ddd; text-align: center; "
                f"font-weight: 600;'>{self.convert_period_to_time(period)}</th>"
            )
        html_table += "</tr></thead><tbody>"

        for day in days_order:
            day_lower = day.lower()
            html_table += (
                "<tr><td style='padding: 12px; border: 1px solid #ddd; "
                "font-weight: bold; background-color: #f9f9f9; width: 120px;'>"
                f"{day}</td>"
            )

            for period in periods_order:
                period_lower = period.lower()
                items = timetable.get(day_lower, {}).get(period_lower, [])
                cell_style = (
                    "padding: 15px; border: 1px solid #ddd; "
                    "vertical-align: middle; min-height: 100px; text-align: center;"
                )

                if items:
                    # Check if reserved
                    is_reserved = any("[RESERVED]" in item for item in items)
                    bg_color = "#e3f2fd" if is_reserved else "#f1f8e9"
                    cell_style += f" background-color: {bg_color};"
                    cell_content = (
                        "<br><hr style='margin: 8px 0; border: 0; "
                        "border-top: 1px solid rgba(0,0,0,0.1);'>"
                    ).join(items)
                else:
                    cell_style += " background-color: #fafafa; color: #999;"
                    cell_content = "<em style='font-size: 0.9em;'>Available</em>"

                html_table += f"<td style='{cell_style}'>{cell_content}</td>"

            html_table += "</tr>"

        html_table += "</tbody></table>"

        return html_table

    def render(self):
        """Render the Timetable View page"""
        _ = st.header("Timetable View")

        if not st.session_state.schedule_generated:
            _ = st.warning(
                "No schedule generated yet. Please schedule courses from the Dashboard."
            )
            return

        # Filters
        all_items = (
            st.session_state.current_schedule + st.session_state.current_reserved
        )
        rooms = ["All"] + extract_unique_values(all_items, "room")
        professors = ["All"] + extract_unique_values(
            st.session_state.current_schedule, "professor"
        )
        years = ["All"] + extract_unique_values(
            st.session_state.current_schedule, "year"
        )

        col1, col2, col3 = st.columns(3)
        with col1:
            selected_room = st.selectbox(
                "Filter by Room", rooms, key="timetable_room_filter"
            )
        with col2:
            selected_professor = st.selectbox(
                "Filter by Professor", professors, key="timetable_professor_filter"
            )
        with col3:
            selected_year = st.selectbox(
                "Filter by Year", years, key="timetable_year_filter"
            )

        # Create timetable data grouped by room
        timetables_by_room = self.create_timetable_data(
            st.session_state.current_schedule,
            st.session_state.current_reserved,
            selected_room,
            selected_professor,
            selected_year,
        )

        if not timetables_by_room:
            _ = st.info("No schedules found matching the selected filters.")
            return

        # Display scheduled courses and reserved slots side by side
        _ = st.markdown("<br>", unsafe_allow_html=True)
        # _ = st.subheader("Scheduled Courses & Reserved Slots")

        # Create two columns for side-by-side display
        col1, col2 = st.columns(2, gap="large")

        # Filter schedule based on selected filters
        filtered_schedule = [
            item
            for item in st.session_state.current_schedule
            if (selected_room == "All" or item.get("room", "") == selected_room)
            and (
                selected_professor == "All"
                or item.get("professor", "") == selected_professor
            )
            and (selected_year == "All" or item.get("year", "") == selected_year)
        ]

        # Filter reserved slots based on selected filters
        filtered_reserved = [
            item
            for item in st.session_state.current_reserved
            if (selected_room == "All" or item.get("room", "") == selected_room)
        ]

        # Left column: Scheduled Courses
        with col1:
            _ = st.markdown("#### Scheduled Courses")
            if filtered_schedule:
                # Create DataFrame with relevant columns
                courses_df = pd.DataFrame(filtered_schedule)

                # Select and rename columns for display
                display_columns = []
                column_mapping = {}

                if "course_name" in courses_df.columns:
                    display_columns.append("course_name")
                    column_mapping["course_name"] = "Course"
                if "year" in courses_df.columns:
                    display_columns.append("year")
                    column_mapping["year"] = "Year"
                if "professor" in courses_df.columns:
                    display_columns.append("professor")
                    column_mapping["professor"] = "Professor"
                if "room" in courses_df.columns:
                    display_columns.append("room")
                    column_mapping["room"] = "Room"
                if "day" in courses_df.columns:
                    display_columns.append("day")
                    column_mapping["day"] = "Day"
                if "period" in courses_df.columns:
                    display_columns.append("period")
                    column_mapping["period"] = "Period"
                if "time_range" in courses_df.columns:
                    display_columns.append("time_range")
                    column_mapping["time_range"] = "Time"

                if display_columns:
                    display_df = courses_df[display_columns].rename(
                        columns=column_mapping
                    )
                    _ = st.dataframe(display_df, width="stretch")
            else:
                _ = st.info("No courses found matching the selected filters.")

        # Right column: Reserved Slots
        with col2:
            _ = st.markdown("#### Reserved Slots")
            if filtered_reserved:
                display_schedule_table(filtered_reserved, "Reserved Slots")
            else:
                _ = st.info("No reserved slots found matching the selected filters.")

        # Display separate table for each room
        _ = st.markdown("<br>", unsafe_allow_html=True)
        _ = st.subheader("Room Timetables")

        # Sort rooms for consistent display
        sorted_rooms = sorted(timetables_by_room.keys())

        for room in sorted_rooms:
            timetable = timetables_by_room[room]

            # Export button for each room
            col1, col2 = st.columns([4, 1])
            with col1:
                _ = st.markdown(
                    f"<h4>Room {room.upper()}</h4>",
                    unsafe_allow_html=True,
                )
            with col2:
                if ui.button(
                    text="📥 Export Excel",
                    key=f"excel_{room}",
                    variant="outline",
                    size="lg",
                    class_name="w-full",
                ):
                    try:
                        excel_data = self.export_timetable_to_excel(room, timetable)
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        _ = st.download_button(
                            label="Download Excel",
                            data=excel_data,
                            file_name=f"timetable_{room}_{timestamp}.xlsx",
                            mime=(
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet"
                            ),
                            key=f"dl_excel_{room}",
                        )
                    except (IOError, ValueError, KeyError) as e:
                        _ = st.error(f"Error exporting to Excel: {str(e)}")

            html_table = self.render_timetable_for_room(room, timetable)
            _ = st.markdown(html_table, unsafe_allow_html=True)
            _ = st.markdown("<br>", unsafe_allow_html=True)
