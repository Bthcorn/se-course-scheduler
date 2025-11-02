"""
Excel Handler for Course Scheduling System
Handles all Excel file operations including loading data and exporting schedules
"""

from typing import Any
import pandas as pd
from pyswip import Prolog
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """Handles all Excel file operations for the course scheduling system"""

    prolog: Prolog

    def __init__(self, prolog: Prolog):
        """Initialize with a Prolog instance for data assertion"""
        self.prolog = prolog

    def load_courses_from_excel(self, excel_file: str) -> None:
        """Load courses from Excel file and assert into Prolog"""
        try:
            df = pd.read_excel(excel_file, sheet_name="Courses")
            for _, row in df.iterrows():
                course_id = str(row["CourseID"])
                course_name = str(row["CourseName"])
                year = int(row["Year"])
                capacity = int(row["RequiredCapacity"])
                # Parse prerequisites (comma-separated)
                prereqs = []
                if pd.notna(row.get("Prerequisites", "")):
                    prereqs = [p.strip() for p in str(row["Prerequisites"]).split(",")]

                # Create Prolog fact
                prereq_list = "[" + ",".join(prereqs) + "]" if prereqs else "[]"
                course_query = (
                    f"assertz(course({course_id}, '{course_name}', {year}, "
                    f"{capacity}, {prereq_list}))"
                )

                _ = list(self.prolog.query(course_query))
            print(f"✓ Loaded {len(df)} courses from Excel")

        except (FileNotFoundError, KeyError, ValueError, Exception) as e:
            print(f"✗ Error loading courses from Excel: {e}")
            raise

    def load_professors_from_excel(self, excel_file: str) -> None:
        """Load professors and their capabilities from Excel"""
        try:
            # Load professors
            df_prof = pd.read_excel(excel_file, sheet_name="Professors")
            for _, row in df_prof.iterrows():
                professor_id = str(row["ProfessorID"])
                professor_name = str(row["ProfessorName"])
                professor_query = (
                    f"assertz(professor({professor_id}, '{professor_name}'))"
                )
                _ = list(self.prolog.query(professor_query))
            # Load teaching capabilities
            df_teach = pd.read_excel(excel_file, sheet_name="CanTeach")
            for _, row in df_teach.iterrows():
                teach_prof_id = str(row["ProfessorID"])
                course_id = str(row["CourseID"])
                teach_query = f"assertz(can_teach({teach_prof_id}, {course_id}))"
                _ = list(self.prolog.query(teach_query))
            # Load preferences (if available)
            if "Preferences" in pd.ExcelFile(excel_file).sheet_names:
                df_pref = pd.read_excel(excel_file, sheet_name="Preferences")
                for _, row in df_pref.iterrows():
                    pref_prof_id = str(row["ProfessorID"])
                    day = str(row["Day"]).lower()
                    period = str(row["Period"]).lower()
                    pref_query = f"assertz(prefers({pref_prof_id}, {day}, {period}))"
                    _ = list(self.prolog.query(pref_query))
            print(f"✓ Loaded {len(df_prof)} professors from Excel")

        except (FileNotFoundError, KeyError, ValueError, Exception) as e:
            print(f"✗ Error loading professors from Excel: {e}")
            raise

    def load_all_data_from_excel(self, excel_file: str) -> None:
        """Load all data (courses and professors) from Excel file"""
        self.load_courses_from_excel(excel_file)
        self.load_professors_from_excel(excel_file)

    def export_schedule_to_excel(
        self,
        schedule: list[dict[str, str]],
        reserved: list[dict[str, str]],
        stats: dict[str, Any],
        output_file: str = "generated_schedule.xlsx",
    ) -> None:
        """Export schedule to Excel file organized by rooms"""
        if not schedule and not reserved:
            print("✗ No schedule to export")
            return

        # Create a writer object
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Get all rooms that have either scheduled or reserved slots
            all_rooms: set[str] = set()
            for item in schedule:
                all_rooms.add(item["room"])
            for item in reserved:
                all_rooms.add(item["room"])

            # Create a summary sheet with room overview
            summary_data = []
            for room in sorted(all_rooms):
                room_schedule = [cls for cls in schedule if cls["room"] == room]
                room_reserved = [res for res in reserved if res["room"] == room]

                summary_data.append(
                    {
                        "Room": room.upper(),
                        "Scheduled Courses": len(room_schedule),
                        "Reserved Slots": len(room_reserved),
                        "Total Activities": len(room_schedule) + len(room_reserved),
                        "Available Slots": 21 - len(room_reserved),
                        "Utilization": (
                            f"{(len(room_schedule) / (21 - len(room_reserved)) * 100):.1f}%"
                            if (21 - len(room_reserved)) > 0
                            else "N/A"
                        ),
                    }
                )

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Room_Summary", index=False)
            # Create detailed room sheets in table format
            for room in sorted(all_rooms):
                # Get all time slots for this room
                room_schedule = [cls for cls in schedule if cls["room"] == room]
                room_reserved = [res for res in reserved if res["room"] == room]

                # Create timetable table format with days as rows and time as columns
                days = [
                    "Monday",
                    "Tuesday",
                    "Wednesday",
                    "Thursday",
                    "Friday",
                    "Saturday",
                    "Sunday",
                ]
                # Period headers with names matching HTML view
                period_configs = [
                    ("Morning (09:00-12:00)", "morning", "09.00-12.00"),
                    ("Afternoon (13:00-16:00)", "afternoon", "13.00-16.00"),
                    ("Evening (17:00-20:00)", "evening", "17.00-20.00"),
                ]

                # Create table data
                table_data = []
                for day in days:
                    day_lower = day.lower()
                    row = {"Day": day}
                    for period_display, period_lower, period_key in period_configs:
                        col_name = period_key

                        # Check if this slot is reserved
                        reserved_slot = next(
                            (
                                r
                                for r in room_reserved
                                if r["day"] == day_lower and r["period"] == period_lower
                            ),
                            None,
                        )
                        if reserved_slot:
                            row[col_name] = f"[RESERVED]\n{reserved_slot['reason']}"
                        else:
                            # Check if this slot has a scheduled course
                            scheduled_slot = next(
                                (
                                    s
                                    for s in room_schedule
                                    if s["day"] == day_lower
                                    and s["period"] == period_lower
                                ),
                                None,
                            )
                            if scheduled_slot:
                                row[col_name] = (
                                    f"{scheduled_slot['course_name']}\n"
                                    f"{scheduled_slot['professor']}"
                                )
                            else:
                                row[col_name] = "[AVAILABLE]"

                    table_data.append(row)
                # Create DataFrame for this room in table format
                room_df = pd.DataFrame(table_data)

                # Write to Excel with proper formatting
                # Write DataFrame starting at row 2 (leaving row 1 for room header)
                room_df.to_excel(
                    writer,
                    sheet_name=f"Room_{room.upper()}",
                    index=False,
                    startrow=1,  # Start at row 2 to leave space for header
                )
                # Get the workbook and worksheet for formatting
                worksheet = writer.sheets[f"Room_{room.upper()}"]

                # Add header row with room name (merged across all columns)
                num_cols = len(period_configs) + 1  # Day column + period columns
                header_col = get_column_letter(num_cols)
                worksheet.merge_cells(f"A1:{header_col}1")
                header_cell = worksheet["A1"]
                header_cell.value = f"Room: {room.upper()}"
                header_cell.font = Font(color="FFFFFF", bold=True, size=14)
                header_cell.fill = PatternFill(
                    start_color="1f77b4", end_color="1f77b4", fill_type="solid"
                )
                header_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                worksheet.row_dimensions[1].height = 30

                # Update column headers to match HTML format (overwrite DataFrame headers)
                headers = ["Day"] + [
                    period_display for period_display, _, _ in period_configs
                ]
                for col_idx, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=2, column=col_idx, value=header)
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(
                        start_color="f0f2f6", end_color="f0f2f6", fill_type="solid"
                    )
                    cell.alignment = Alignment(
                        horizontal="center" if col_idx > 1 else "left",
                        vertical="center",
                    )
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                # Adjust column widths for better readability
                worksheet.column_dimensions["A"].width = 15  # Day column
                worksheet.column_dimensions["B"].width = 35  # Morning column
                worksheet.column_dimensions["C"].width = 35  # Afternoon column
                worksheet.column_dimensions["D"].width = 35  # Evening column

            # Add statistics sheets
            # Professor workload
            if stats.get("professor_workload"):
                workload_data = stats["professor_workload"]
                if isinstance(workload_data, list):
                    workload_df = pd.DataFrame(workload_data)
                    workload_df.to_excel(
                        writer, sheet_name="Professor_Workload", index=False
                    )

            # Detailed room utilization
            if stats.get("room_utilization"):
                util_data = stats["room_utilization"]
                if isinstance(util_data, list):
                    util_df = pd.DataFrame(util_data)

                    # Add reserved slots info to utilization
                    reserved_slots_list: list[int] = []
                    available_slots_list: list[int] = []

                    for room_stat in util_data:
                        room_stat_dict = dict(room_stat)
                        room_name = str(room_stat_dict["room"])
                        reserved_count = len(
                            [r for r in reserved if r["room"] == room_name]
                        )
                        reserved_slots_list.append(reserved_count)
                        available_slots_list.append(
                            int(room_stat_dict["total_slots"]) - reserved_count
                        )

                    util_df["reserved_slots"] = reserved_slots_list
                    util_df["available_slots"] = available_slots_list

                    util_df.to_excel(writer, sheet_name="Room_Utilization", index=False)
        print(f"✓ Schedule exported to {output_file}")

    def export_schedule_dataframe(
        self, schedule: list[dict[str, str]], reserved: list[dict[str, str]]
    ) -> pd.DataFrame:
        """Convert schedule data to a pandas DataFrame for easy manipulation"""
        all_data = []

        # Add scheduled courses
        for item in schedule:
            all_data.append(
                {
                    "Type": "Course",
                    "Room": item["room"],
                    "Day": item["day"],
                    "Period": item["period"],
                    "Time Range": item["time_range"],
                    "Activity": item["course_name"],
                    "Person": item["professor"],
                    "Course ID": item["course_id"],
                }
            )

        # Add reserved slots
        for item in reserved:
            all_data.append(
                {
                    "Type": "Reserved",
                    "Room": item["room"],
                    "Day": item["day"],
                    "Period": item["period"],
                    "Time Range": item["time_range"],
                    "Activity": f"[RESERVED] {item['reason']}",
                    "Person": "N/A",
                    "Course ID": "N/A",
                }
            )

        return pd.DataFrame(all_data)
